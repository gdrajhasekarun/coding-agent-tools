import os
import json
import openpyxl

def main():
    try:
        # Load previous context if available
        prev = json.loads(sys.stdin.read()) if not sys.stdin.isatty() else {}
        
        # Get the path of the uploaded Excel file
        excel_file_path = os.environ.get("UPLOADED_FILE_PATH")
        if not excel_file_path or not os.path.exists(excel_file_path):
            raise FileNotFoundError("Uploaded Excel file does not exist.")

        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_file_path)
        results = {}

        # Iterate through the sheets in the workbook
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            # Store the data for each sheet in results
            results[sheet_name] = data

        # Define the output path for saving intermediate results
        work_dir = os.environ.get("WORK_DIR")
        if not work_dir or not os.path.isdir(work_dir):
            raise FileNotFoundError("Working directory does not exist.")

        # Write results to JSON file
        output_file_path = os.path.join(work_dir, "excel_data.json")
        with open(output_file_path, 'w') as json_file:
            json.dump(results, json_file)

        # Prepare the summary
        summary = {
            "message": "Excel file processed successfully.",
            "sheets": list(results.keys()),
            "output_file": output_file_path
        }
        
        print(json.dumps(summary))

    except Exception as e:
        print(json.dumps({"error": str(e)}))


if __name__ == "__main__":
    import sys
    main()

# METADATA: {"description": "Reads an Excel file and outputs its content as JSON.", "inputs": ["UPLOADED_FILE_PATH"], "outputs": ["JSON summary of sheet content"], "limitations": "Only processes Excel files; does not validate contents."}